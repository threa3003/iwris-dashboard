"""
process_data.py
----------------
Merges:
  1. Groundwater state files  (e.g. Andhra_Pradesh.xlsx)
     - sheet "<State>_Stn"  -> hierarchical station metadata (has Data Status etc.)
     - sheet "<State>_Data" -> tidy long-format readings (Date, Water Level (m))
  2. Rainfall per-station files (e.g. Rainfall_Alladupalli.xlsx)
     - sheet "Metadata - <name>" -> key/value station metadata
     - sheet "Rainfall - <name>" -> tidy long-format readings (Data Time, Data Value)

Output (written to ./output/):
  stations.json  -> one row per station, both types, common schema
  readings.json  -> one row per (station_code, date, parameter, value)

Folder layout expected on disk:

  data-raw/
    groundwater/
      Andhra_Pradesh.xlsx
      Bihar.xlsx
      ...
    rainfall/
      Andhra_Pradesh/
        Rainfall_Alladupalli.xlsx
        Rainfall_SomeOtherStation.xlsx
      Bihar/
        ...

Adjust RAW_DIR below if your folder names differ.
"""

import json
import re
import multiprocessing as mp
from pathlib import Path
from datetime import datetime

import openpyxl

RAW_DIR = Path("data-raw")
GW_DIR = RAW_DIR / "groundwater"
RAIN_DIR = RAW_DIR / "rainfall"
OUT_DIR = Path("output")
OUT_DIR.mkdir(exist_ok=True)

stations = {}   # station_code -> station dict
readings = []   # list of {station_code, date, parameter, value, unit}


def to_iso_date(value):
    """Normalize whatever date format shows up into 'YYYY-MM-DD'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.upper() == "NO DATA":
        return None
    # handles 'YYYY-MM-DDTHH:MM:SS' and plain 'YYYY-MM-DD'
    return s[:10]


def safe_float(value):
    try:
        if value is None or str(value).strip().upper() == "NO DATA":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# 1. GROUNDWATER FILES
# ---------------------------------------------------------------------------

def process_groundwater_file(path: Path, readings_file):
    state_guess = path.stem.replace("_", " ")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  !! SKIPPED (could not open): {path.name} -> {e}")
        return

    stn_sheet = next((s for s in wb.sheetnames if s.endswith("_Stn")), None)
    data_sheet = next((s for s in wb.sheetnames if s.endswith("_Data")), None)

    # --- metadata from the hierarchical "_Stn" sheet ---
    if stn_sheet:
        ws = wb[stn_sheet]
        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if header:
            col = {name: i for i, name in enumerate(header) if name}

            current_district = current_tehsil = current_block = None
            consecutive_blank = 0
            BLANK_LIMIT = 500  # IWRIS files can report a phantom used-range far
            # beyond their real data (formatting applied to empty cells), which
            # makes iter_rows() try to loop through millions of blank rows and
            # hang. Real data is always contiguous, so once we see this many
            # blank rows in a row we've clearly hit the phantom tail — stop.

            for row in row_iter:
                level = row[0] if row else None
                if level is None:
                    consecutive_blank += 1
                    if consecutive_blank >= BLANK_LIMIT:
                        break
                    continue
                consecutive_blank = 0
                level = str(level).strip()
                if level == "DISTRICT":
                    current_district = row[1]
                elif level == "TEHSIL":
                    current_tehsil = row[1]
                elif level == "BLOCK":
                    current_block = row[1]
                elif level == "STATION":
                    code = row[col.get("Station Code", 2)]
                    if not code:
                        continue
                    code = str(code).strip()
                    stations[code] = {
                        "station_code": code,
                        "station_name": row[col.get("Station Name", 3)],
                        "parameter": "groundwater",
                        "state": state_guess,
                        "district": current_district,
                        "tehsil": current_tehsil,
                        "block": current_block,
                        "latitude": row[col.get("Latitude")],
                        "longitude": row[col.get("Longitude")],
                        "agency": row[col.get("Agency")],
                        "well_type": row[col.get("Type of Well")],
                        "aquifer_type": row[col.get("Aquifer Type")],
                        "well_depth": row[col.get("Well Depth")],
                        "data_status": row[col.get("Data Status")],
                    }

    # --- readings from the tidy "_Data" sheet ---
    if data_sheet:
        ws = wb[data_sheet]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            col = {name: i for i, name in enumerate(header) if name}
            consecutive_blank = 0
            BLANK_LIMIT = 500
            row_count = 0

            for row in rows:
                code = row[col.get("Station Code", 4)] if row else None
                if code is None:
                    consecutive_blank += 1
                    if consecutive_blank >= BLANK_LIMIT:
                        break
                    continue
                consecutive_blank = 0
                code = str(code).strip()
                date = to_iso_date(row[col.get("Date")])
                value = safe_float(row[col.get("Water Level (m)")])
                if date is None or value is None:
                    continue  # skip "NO DATA" rows
                readings_file.write(json.dumps({
                    "station_code": code,
                    "date": date,
                    "parameter": "groundwater_level_m",
                    "value": value,
                }, default=str) + "\n")
                row_count += 1
                if row_count % 50000 == 0:
                    print(f"    ...{path.name}: {row_count} rows written so far", flush=True)

    wb.close()


# ---------------------------------------------------------------------------
# 2. RAINFALL FILES (one file per station)
# ---------------------------------------------------------------------------

def process_rainfall_file(path: Path, state_from_folder: str, readings_file):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  !! SKIPPED (could not open): {path.name} -> {e}")
        return

    meta_sheet = next((s for s in wb.sheetnames if s.startswith("Metadata")), None)
    data_sheet = next((s for s in wb.sheetnames if s.startswith("Rainfall -")), None)

    meta = {}
    if meta_sheet:
        ws = wb[meta_sheet]
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                meta[str(row[0]).strip()] = row[1]

    code = meta.get("Station Code")
    if not code:
        wb.close()
        return
    code = str(code).strip()

    stations[code] = {
        "station_code": code,
        "station_name": meta.get("Station Name"),
        "parameter": "rainfall",
        "state": meta.get("State", state_from_folder),
        "district": meta.get("District"),
        "tehsil": meta.get("Tehsil"),
        "block": None,
        "latitude": meta.get("Latitude"),
        "longitude": meta.get("Longitude"),
        "agency": meta.get("Agency Name"),
        "well_type": None,
        "aquifer_type": None,
        "well_depth": None,
        "data_status": meta.get("Station Status"),
    }

    if data_sheet:
        ws = wb[data_sheet]
        rows = ws.iter_rows(values_only=True)
        # skip the ~6 preamble rows, find the real header row
        header = None
        for row in rows:
            if row and row[0] == "Data Type Code":
                header = row
                break
        if header:
            col = {name: i for i, name in enumerate(header) if name}
            for row in rows:
                if not row or row[col.get("Data Time", 2)] is None:
                    continue
                date = to_iso_date(row[col.get("Data Time")])
                value = safe_float(row[col.get("Data Value")])
                if date is None or value is None:
                    continue
                readings_file.write(json.dumps({
                    "station_code": code,
                    "date": date,
                    "parameter": "rainfall_mm",
                    "value": value,
                }, default=str) + "\n")

    wb.close()


# ---------------------------------------------------------------------------
# TIMEOUT PROTECTION
# ---------------------------------------------------------------------------
# Some corrupted xlsx files cause openpyxl to hang instead of raising a clean
# error. Running each file's processing in a short-lived subprocess with a
# hard timeout means one bad file can never freeze the whole batch.
#
# Readings are streamed straight to readings.jsonl (one JSON object per line,
# appended as we go) instead of being held in a Python list. This keeps
# memory flat regardless of how many files/readings we process — fixes the
# multi-GB RAM growth seen when everything was kept in memory.

FILE_TIMEOUT_SECONDS = 30
READINGS_PATH = None  # set in main()


def _run_gw_worker(path_str, readings_path_str, queue):
    with open(readings_path_str, "a", encoding="utf-8") as rf:
        process_groundwater_file(Path(path_str), rf)
    queue.put(dict(stations))


def _run_rain_worker(path_str, state_name, readings_path_str, queue):
    with open(readings_path_str, "a", encoding="utf-8") as rf:
        process_rainfall_file(Path(path_str), state_name, rf)
    queue.put(dict(stations))


def process_with_timeout(worker_fn, args):
    queue = mp.Queue()
    p = mp.Process(target=worker_fn, args=(*args, queue))
    p.start()
    p.join(FILE_TIMEOUT_SECONDS)

    if p.is_alive():
        p.terminate()
        p.join()
        return None, "TIMED OUT (hung file, skipped)"

    if not queue.empty():
        result_stations = queue.get()
        return result_stations, None

    return None, f"exited with code {p.exitcode} (crashed, skipped)"


def save_stations_checkpoint():
    (OUT_DIR / "stations.json").write_text(
        json.dumps(list(stations.values()), indent=2, default=str)
    )


# ---------------------------------------------------------------------------
# RUN
# ---------------------------------------------------------------------------

def main():
    readings_path = OUT_DIR / "readings.jsonl"
    # start fresh each run so re-runs don't duplicate lines
    readings_path.write_text("")

    gw_files = sorted(
        f for f in (GW_DIR.glob("*.xlsx") if GW_DIR.exists() else [])
        if not f.name.startswith("~$")
    )
    print(f"Found {len(gw_files)} groundwater state files")
    skipped = []

    for i, f in enumerate(gw_files, 1):
        print(f"  [{i}/{len(gw_files)}] processing GW: {f.name}", flush=True)
        result_stations, err = process_with_timeout(
            _run_gw_worker, (str(f), str(readings_path))
        )
        if err:
            print(f"    !! SKIPPED: {f.name} -> {err}", flush=True)
            skipped.append(str(f))
        else:
            stations.update(result_stations)
        if i % 5 == 0:
            save_stations_checkpoint()

    rain_files = []
    if RAIN_DIR.exists():
        for state_folder in RAIN_DIR.iterdir():
            if state_folder.is_dir():
                for f in state_folder.glob("*.xlsx"):
                    if not f.name.startswith("~$"):
                        rain_files.append((f, state_folder.name.replace("_", " ")))
    print(f"Found {len(rain_files)} rainfall station files")

    for i, (f, state_name) in enumerate(rain_files, 1):
        print(f"  [{i}/{len(rain_files)}] processing Rainfall: {f.name}", flush=True)
        result_stations, err = process_with_timeout(
            _run_rain_worker, (str(f), state_name, str(readings_path))
        )
        if err:
            print(f"    !! SKIPPED: {f.name} -> {err}", flush=True)
            skipped.append(str(f))
        else:
            stations.update(result_stations)
        if i % 20 == 0:
            save_stations_checkpoint()
            line_count = sum(1 for _ in open(readings_path, encoding="utf-8"))
            print(f"    [checkpoint: {len(stations)} stations, {line_count} readings so far]", flush=True)

    save_stations_checkpoint()
    total_readings = sum(1 for _ in open(readings_path, encoding="utf-8"))

    print(f"\nDone. {len(stations)} stations, {total_readings} readings.")
    print(f"Wrote {OUT_DIR / 'stations.json'} and {readings_path}")
    print("(readings.jsonl is one JSON object per line — convert to a single")
    print(" JSON array later if your frontend needs that, or read it")
    print(" line-by-line, which is more memory-efficient anyway.)")
    if skipped:
        print(f"\n{len(skipped)} files were skipped (hung or crashed):")
        for s in skipped:
            print(" -", s)


if __name__ == "__main__":
    main()
