"""
diagnose_hang.py
-----------------
Opens ONE specific file step-by-step with print statements after every
major operation, so we can see exactly which step never completes.
Run this directly (no timeout wrapper) and watch which line prints last.
"""

import sys
import time
from pathlib import Path

TARGET_FILE = r"data-raw\groundwater\Andhra Pradesh.xlsx"

print(f"Target: {TARGET_FILE}", flush=True)
path = Path(TARGET_FILE)
print(f"Exists: {path.exists()}", flush=True)
print(f"Size: {path.stat().st_size / 1_000_000:.1f} MB", flush=True)

print("\n[1] Importing openpyxl...", flush=True)
t0 = time.time()
import openpyxl
print(f"    done ({time.time()-t0:.1f}s)", flush=True)

print("\n[2] Opening workbook (read_only=True, data_only=True)...", flush=True)
t0 = time.time()
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"    done ({time.time()-t0:.1f}s)", flush=True)

print("\n[3] Listing sheet names...", flush=True)
t0 = time.time()
names = wb.sheetnames
print(f"    done ({time.time()-t0:.1f}s): {names}", flush=True)

stn_sheet = next((s for s in names if s.endswith("_Stn")), None)
data_sheet = next((s for s in names if s.endswith("_Data")), None)
print(f"\n    stn_sheet={stn_sheet}, data_sheet={data_sheet}", flush=True)

if stn_sheet:
    print(f"\n[4] Opening _Stn sheet '{stn_sheet}'...", flush=True)
    t0 = time.time()
    ws = wb[stn_sheet]
    print(f"    done ({time.time()-t0:.1f}s)", flush=True)

    print(f"\n[5] Checking ws.max_row / max_column (reported used range)...", flush=True)
    t0 = time.time()
    print(f"    max_row={ws.max_row}, max_column={ws.max_column} ({time.time()-t0:.1f}s)", flush=True)

    print(f"\n[6] Getting row iterator and reading first row...", flush=True)
    t0 = time.time()
    row_iter = ws.iter_rows(values_only=True)
    first_row = next(row_iter, None)
    print(f"    done ({time.time()-t0:.1f}s): {first_row}", flush=True)

    print(f"\n[7] Reading ALL remaining rows, timing each, reporting slow ones...", flush=True)
    row_num = 1
    while True:
        t0 = time.time()
        row = next(row_iter, None)
        elapsed = time.time() - t0
        row_num += 1
        if row is None:
            print(f"    reached end of sheet at row {row_num}", flush=True)
            break
        if elapsed > 0.5:
            print(f"    !! SLOW row {row_num}: {elapsed:.2f}s -> {row}", flush=True)
        if row_num % 1000 == 0:
            print(f"    ...at row {row_num} ({elapsed:.3f}s for this row)", flush=True)
        if row_num > 200000:
            print(f"    stopping scan at 200,000 rows as a safety cap", flush=True)
            break

if data_sheet:
    print(f"\n[8] Opening _Data sheet '{data_sheet}'...", flush=True)
    t0 = time.time()
    ws2 = wb[data_sheet]
    print(f"    done ({time.time()-t0:.1f}s)", flush=True)

    print(f"\n[9] Checking ws2.max_row / max_column...", flush=True)
    t0 = time.time()
    print(f"    max_row={ws2.max_row}, max_column={ws2.max_column} ({time.time()-t0:.1f}s)", flush=True)

    print(f"\n[10] Reading ALL remaining rows of _Data sheet, timing each...", flush=True)
    row_iter2 = ws2.iter_rows(values_only=True)
    row_num2 = 1
    while True:
        t0 = time.time()
        row = next(row_iter2, None)
        elapsed = time.time() - t0
        row_num2 += 1
        if row is None:
            print(f"    reached end of sheet at row {row_num2}", flush=True)
            break
        if elapsed > 0.5:
            print(f"    !! SLOW row {row_num2}: {elapsed:.2f}s -> {row}", flush=True)
        if row_num2 % 1000 == 0:
            print(f"    ...at row {row_num2} ({elapsed:.3f}s for this row)", flush=True)
        if row_num2 > 200000:
            print(f"    stopping scan at 200,000 rows as a safety cap", flush=True)
            break

print("\nDIAGNOSTIC COMPLETE — if you see this, nothing hung in the first 20 rows.", flush=True)
